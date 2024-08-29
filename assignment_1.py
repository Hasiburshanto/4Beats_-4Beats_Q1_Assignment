import os
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
import random
from datetime import datetime

def install_packages():
    os.system('pip3 install -U selenium')
    os.system('pip3 install openpyxl')

def google_search(query):
    chrome_driver = webdriver.Chrome()

    try:
        chrome_driver.get("https://www.google.com")
        chrome_search_box = chrome_driver.find_element(By.CLASS_NAME, "gLFyf")
        chrome_search_box.send_keys(query)
        time.sleep(2)
        options = chrome_driver.find_elements(By.CLASS_NAME, "G43f7e")
        all_suggestions = []
        for option_1 in options:
            for line in option_1.text.splitlines():
                all_suggestions.append(line)
        longest_option = max(all_suggestions, key=len)
        shortest_option = min(all_suggestions, key=len)

        return longest_option, shortest_option
    finally:
        chrome_driver.quit()

def update_excel_file(file_path):
    current_day = datetime.now().strftime('%A')
    print(current_day)
    input_file = openpyxl.load_workbook(file_path)
    current_day_sheet = input_file[current_day]

    for row in current_day_sheet.iter_rows(min_row=1, max_row=current_day_sheet.max_row, min_col=1, max_col=current_day_sheet.max_column):
        row_str = ''
        for cell in row:
            if cell.value is None:
                row_str += 'None '
            else:
                row_str += str(cell.value) + ' '

        if "Keyword" in row_str:
            longest, shortest = google_search(row[2].value)
            row[3].value = longest
            row[4].value = shortest

    input_file.save(file_path)
    print("File overwritten successfully!")

if __name__ == "__main__":
    file_path = '/Users/shanto/Desktop/4beats/4BeatsQ1.xlsx'  # Specify the file path here
    install_packages()
    update_excel_file(file_path)
